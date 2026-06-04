#!/usr/bin/env python3
"""Regenerate HOD menu data from XLSX into prod index.html, admin.html, hod-menu.ts.
Tax model (matches printed bills):
  SC  = 10% * subtotal_all (food + ALL drinks, alcoholic and non)
  GST = 5%  * (food + non-alcoholic drinks + SC)   [alcohol exempt from GST]
  CGST = SGST = GST/2 displayed separately on customer breakdown
"""
import openpyxl, json, re, sys, glob, os

# --- Locate newest XLSX inputs ---
def newest(pattern):
    matches = sorted(glob.glob(pattern))
    if not matches: raise SystemExit(f'No file matches {pattern}')
    return matches[-1]

FOOD_XLSX = newest('attached_assets/HOD_FOOD_MENU_*.xlsx')
BAR_XLSX  = newest('attached_assets/SATISH_NEW_BAR_MENU_*.xlsx')
print(f'food: {FOOD_XLSX}', file=sys.stderr)
print(f'bar : {BAR_XLSX}', file=sys.stderr)

# --- Title-case helper ---
def tcase(s):
    s = re.sub(r'\s+', ' ', str(s)).strip()
    s = s.title()
    s = s.replace(' &Amp; ', ' & ').replace("'S", "'s")
    return s

# === FOOD parser ===
# Food XLSX layout: col0=VEG/NONVEG, col2=name OR cat header, col3=oldPrice, col4=PRICE
FOOD_CAT_HEADERS = {
    'SOUP':'Soups','SALAD':'Salads','BAR BITES':'Bar Bites','GALLI KI CHAT':'Chaat',
    'CHARGRILLED APPITIZER':'Chargrilled','PLATTERS':'Platters','ORIENTAL APPITIZER':'Oriental',
    'BAO':'Bao','DIMSUM':'Dimsum','INTERNATIONAL APPITIZERS':'International Starters',
    'PASTA & PIZZA':'Pasta','PIZZA':'Pizza','INTERNATION MAINS':'International Mains',
    'COASTAL':'Coastal','MAIN COURSE':'Main Course','BIRYANI AND RICE':'Biryani & Rice',
    'FRIED RICE AND NOODLES':'Rice & Noodles','BREAD':'Breads','DESSERTS':'Desserts',
}
GROUP_FOR_FOODCAT = {c:'food' for c in FOOD_CAT_HEADERS.values()}
SLUG_FOR_FOODCAT = {c: 'food-' + re.sub(r'[^a-z0-9]+','-',c.lower()).strip('-') for c in FOOD_CAT_HEADERS.values()}

def parse_food():
    wb = openpyxl.load_workbook(FOOD_XLSX, data_only=True)
    ws = wb['Sheet1']
    cats = []  # list of {cat, items:[]}
    cur = None
    last_veg = None
    for row in ws.iter_rows(values_only=True):
        veg, _, name, oldp, price, _loc = row[:6]
        if name is None: continue
        name_s = str(name).strip()
        if not name_s: continue
        if veg: last_veg = str(veg).strip().upper().startswith('VEG') and 'NON' not in str(veg).strip().upper()
        # Detect category header: only name set, no prices/veg
        upper = name_s.upper()
        if upper in FOOD_CAT_HEADERS and price is None and oldp is None:
            label = FOOD_CAT_HEADERS[upper]
            cur = {'cat': label, 'items': []}
            cats.append(cur)
            continue
        # Skip description rows (no prices, no veg)
        if not isinstance(price, (int, float)) or price <= 0:
            continue
        if cur is None:
            cur = {'cat': 'Misc', 'items': []}
            cats.append(cur)
        # Heuristic veg
        is_veg = True
        if veg is None:
            is_veg = last_veg if last_veg is not None else True
        else:
            v = str(veg).strip().upper()
            is_veg = v.startswith('VEG') and 'NON' not in v
        item = {'n': tcase(name_s), 'p': int(price), 't': 'food', 'alc': False, 'v': bool(is_veg)}
        cur['items'].append(item)
    # Drop empty cats
    cats = [c for c in cats if c['items']]
    return cats

# === BAR parser ===
# Bar XLSX layout: data in col5 (name), col6 (price/glass), col7 (bottle price)
# Category headers detected by col6 like '30 ML', 'BOTTLE', 'GLASS', '275 ML', '330 ML', '500 ML'
BAR_CATEGORY_DEFS = [
    # (header_match_substr, label, group, alcoholic)
    ('FRESH CRAFT BEER',         'Fresh Craft Beer',     'beer-wine',  True),
    ('BREEZERS',                 'Breezers',             'beer-wine',  True),
    ('BOTTLE BEER',              'Bottle Beer',          'beer-wine',  True),
    ('CAN BEER 330',             'Can Beer (330ml)',     'beer-wine',  True),
    ('CAN BEER 500',             'Can Beer (500ml)',     'beer-wine',  True),
    ('SINGLE MALT',              'Single Malt',          'spirits',    True),
    ('AMERICAN / IRISH WHISKY',  'American/Irish Whisky','spirits',    True),
    ('SCOTCH / IMFL',            'Scotch & IMFL Whisky', 'spirits',    True),
    ('VODKA',                    'Vodka',                'spirits',    True),
    ('RUM',                      'Rum',                  'spirits',    True),
    ('GIN',                      'Gin',                  'spirits',    True),
    ('COGNAC',                   'Cognac & Brandy',      'spirits',    True),
    ('TEQUILA',                  'Tequila',              'spirits',    True),
    ('LIQUOR',                   'Liqueurs',             'spirits',    True),
    ('SPARKLING WINES',          'Sparkling Wines',      'beer-wine',  True),
    ('IMPORTED AN DOMESTIC',     'Wines',                'beer-wine',  True),
    ('SIGNATURE MOCKTAILS',      'Signature Mocktails',  'soft',       False),
    ('TALL AND HANDSOME',        'Tall & Handsome',      'cocktails',  True),
    ('WORLDWIDE COCKTAILS',      'Cocktails',            'cocktails',  True),
    ('SHOOTERS',                 'Shooters',             'cocktails',  True),
    ('FLAMERS',                  'Flamers',              'cocktails',  True),
    ('MOCKTAILS',                'Mocktails',            'soft',       False),  # standalone
    ('AREATED BEVERAGE',         'Soft Drinks',          'soft',       False),
]
def find_bar_cat(header):
    h = header.upper()
    for k, label, group, alc in BAR_CATEGORY_DEFS:
        if k in h: return label, group, alc
    return None

def split_multi(name, price):
    """Split items like 'TOIT TINT WIT (330ML/500ML/PITCHER/TOWER)' price '402/517/1380/2530'
       or 'BACARDI WHITE / LEMON / ORANGE' price '370/370/370'."""
    if not isinstance(price, str) or '/' not in price:
        return [(name, price)]
    prices = [p.strip() for p in price.split('/')]
    try:
        pints = [int(p) for p in prices]
    except ValueError:
        return [(name, price)]
    # Pattern A: name has '(...) ' with slash variants
    m = re.search(r'\(([^)]+)\)', name)
    if m and '/' in m.group(1):
        variants = [v.strip() for v in m.group(1).split('/')]
        if len(variants) == len(pints):
            base = name.replace(m.group(0), '').strip()
            return [(f'{base} ({v})', pints[i]) for i, v in enumerate(variants)]
    # Pattern B: name has ' / ' separators
    if ' / ' in name or '/' in name:
        parts = re.split(r'\s*/\s*', name)
        # strip common prefix from variants except first
        if len(parts) == len(pints) and len(parts) > 1:
            head = parts[0].strip()
            head_words = head.split()
            if len(head_words) >= 2:
                prefix = ' '.join(head_words[:-1])
                last_first = head_words[-1]
                variants = [last_first] + [p.strip() for p in parts[1:]]
                return [(f'{prefix} {v}'.strip(), pints[i]) for i, v in enumerate(variants)]
            return [(p.strip(), pints[i]) for i, p in enumerate(parts)]
    # fallback: take first price only
    return [(name, pints[0])]

def parse_bar():
    wb = openpyxl.load_workbook(BAR_XLSX, data_only=True)
    ws = wb['BAR MENU']
    cats = []
    cur = None
    cur_alc = True
    for row in ws.iter_rows(values_only=True):
        cells = list(row[:8])
        # Bar data lives in cols 5,6,7
        c5, c6, c7 = cells[5], cells[6], cells[7]
        if c5 is None: continue
        c5s = str(c5).strip()
        if not c5s: continue
        # Header detection: col6 is text like 'GLASS','BOTTLE','30 ML','275 ML','330 ML','500 ML'
        # OR row has only c5 set (a section title we recognize)
        is_header = False
        cat_match = find_bar_cat(c5s)
        if cat_match:
            # Treat as header IF it matches our known list AND price col is non-numeric
            if c6 is None or isinstance(c6, str):
                is_header = True
        if is_header:
            label, group, alc = cat_match
            cur = {'cat': label, 'items': [], 'group': group, 'alc': alc}
            cur_alc = alc
            cats.append(cur)
            continue
        if cur is None:
            cur = {'cat':'Misc','items':[],'group':'spirits','alc':True}
            cats.append(cur)
        # Item row: c5=name, c6=price (glass/30ml), c7=bottle price (optional)
        # Skip rows with no usable price
        if c6 is None and c7 is None: continue
        # Strip {ingredients...} from cocktail names
        name = re.sub(r'\s*\{[^}]*\}', '', c5s).strip()
        # 30ml / Bottle pair (spirits): create two items
        if isinstance(c6, (int, float)) and isinstance(c7, (int, float)) and cur['group'] == 'spirits':
            cur['items'].append({'n': tcase(name) + ' (30ml)', 'p': int(c6), 't':'drink', 'alc': cur_alc})
            cur['items'].append({'n': tcase(name) + ' (Bottle)','p': int(c7), 't':'drink', 'alc': cur_alc})
            continue
        if isinstance(c6, (int, float)) and isinstance(c7, (int, float)):
            # glass + bottle (wine)
            cur['items'].append({'n': tcase(name) + ' (Glass)', 'p': int(c6), 't':'drink', 'alc': cur_alc})
            cur['items'].append({'n': tcase(name) + ' (Bottle)','p': int(c7), 't':'drink', 'alc': cur_alc})
            continue
        # Multi-variant string price
        price = c6 if c6 is not None else c7
        for sub_name, sub_price in split_multi(name, price):
            if not isinstance(sub_price, (int, float)): continue
            cur['items'].append({'n': tcase(sub_name), 'p': int(sub_price), 't':'drink', 'alc': cur_alc})
    # Strip group/alc keys from cat dict for JS export (keep only cat+items)
    out = []
    for c in cats:
        if not c['items']: continue
        out.append({'cat': c['cat'], 'items': c['items'], '_group': c['group'], '_alc': c['alc']})
    return out

food = parse_food()
bar  = parse_bar()
print(f'parsed: {sum(len(c["items"]) for c in food)} food in {len(food)} cats; {sum(len(c["items"]) for c in bar)} bar in {len(bar)} cats', file=sys.stderr)

# === Output 1: prod index.html JS vars (single line) ===
# Strip _group/_alc from bar before serializing for prod JS (keep alc & t on items only)
bar_for_js = [{'cat': c['cat'], 'items': c['items']} for c in bar]
food_js_line = 'var HOD_FOOD_MENU=' + json.dumps(food, separators=(',',':')) + ';'
bar_js_line  = 'var HOD_BAR_MENU='  + json.dumps(bar_for_js, separators=(',',':')) + ';'
os.makedirs('/tmp/menu_out', exist_ok=True)
with open('/tmp/menu_out/food_var.js','w') as f: f.write(food_js_line)
with open('/tmp/menu_out/bar_var.js','w')  as f: f.write(bar_js_line)

# === Output 2: TS MenuItem[] for hod-menu.ts ===
group_label_map = {
    'spirits':'SPIRITS','beer-wine':'BEER & WINE','cocktails':'COCKTAILS','soft':'SOFT DRINKS','food':'FOOD',
}
items_ts = []
nid = 0
for c in food:
    slug = SLUG_FOR_FOODCAT.get(c['cat'], 'food-' + re.sub(r'[^a-z0-9]+','-',c['cat'].lower()).strip('-'))
    for it in c['items']:
        nid += 1
        items_ts.append({
            'id': f'hod{nid}','name': it['n'],'category': slug,'group':'food',
            'price': it['p'],'isAlcohol': False,'available': True,'isVeg': it['v'],
        })
for c in bar:
    slug = re.sub(r'[^a-z0-9]+','-', c['cat'].lower()).strip('-')
    slug = f"{c['_group']}-{slug}"
    for it in c['items']:
        nid += 1
        items_ts.append({
            'id': f'hod{nid}','name': it['n'],'category': slug,'group': c['_group'],
            'price': it['p'],'isAlcohol': bool(c['_alc']),'available': True,
        })

ts_lines = []
ts_lines.append('// AUTO-GENERATED by scripts/src/regen-hod-from-xlsx.py')
ts_lines.append('// Source XLSX: ' + os.path.basename(FOOD_XLSX) + ' + ' + os.path.basename(BAR_XLSX))
ts_lines.append('// Tax model: SC 10% on ALL items; GST 5% on (food + non-alcoholic + SC); alcohol exempt from GST.')
ts_lines.append('import type { MenuItem } from "./types";')
ts_lines.append('')
# Build category labels from used slugs
cat_labels = {}
for c in food:
    slug = SLUG_FOR_FOODCAT.get(c['cat'], 'food-' + re.sub(r'[^a-z0-9]+','-',c['cat'].lower()).strip('-'))
    cat_labels[slug] = c['cat'].upper()
for c in bar:
    slug = f"{c['_group']}-" + re.sub(r'[^a-z0-9]+','-', c['cat'].lower()).strip('-')
    cat_labels[slug] = c['cat'].upper()
ts_lines.append('export const HOD_CATEGORY_LABELS: Record<string, string> = ' + json.dumps(cat_labels, indent=2) + ';')
ts_lines.append('')
ts_lines.append('export const HOD_GROUP_ORDER = ["spirits","beer-wine","cocktails","soft","food"] as const;')
ts_lines.append('export const HOD_GROUP_LABELS: Record<string,string> = ' + json.dumps(group_label_map, indent=2) + ';')
ts_lines.append('')
ts_lines.append('export const HOD_MENU_ITEMS: MenuItem[] = [')
for it in items_ts:
    ts_lines.append('  ' + json.dumps(it) + ',')
ts_lines.append('];')
with open('artifacts/pos-system/src/lib/hod-menu.ts','w') as f:
    f.write('\n'.join(ts_lines) + '\n')

print(f'WROTE: hod-menu.ts ({len(items_ts)} items)', file=sys.stderr)
print(f'WROTE: /tmp/menu_out/food_var.js, bar_var.js (single-line JS vars)', file=sys.stderr)

# Quick price spot-check
checks = [('Mushroom Cappuccino Soup', 228), ('Tomato Basil Soup', 185), ('Salted French Fries', 235),
          ('Kingfisher Ultra', None), ('Red Bull', 350), ('Coke', 100)]
print('\n=== PRICE CHECKS ===', file=sys.stderr)
for nm, exp in checks:
    matches = [it for it in items_ts if nm.lower() in it['name'].lower()]
    for m in matches[:2]:
        ok = exp is None or m['price'] == exp
        print(f'  {"OK " if ok else "!! "}{m["name"]}: ₹{m["price"]} (expected {exp})', file=sys.stderr)
